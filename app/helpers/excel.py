import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from typing import Any
from datetime import datetime
from app.helpers.value import is_empty, random_string, format_snakecase
from app.config import file_dir

def create_excel(column_data: dict[str, str], row_data: list[dict[str, Any]], file_name: str | None = None, sub_path: str | None = None) -> dict[str, Any]:
    if isinstance(file_name, str) and not is_empty(file_name):
        file_name = file_name.split(".")[0].strip()
    else:
        file_name = "output"

    # sanitize file name
    file_name = re.sub(r"[^a-zA-Z0-9 _.\-]", "", str(file_name)).strip().replace("  ", " ").replace(" ", "_")
    # concat with random string and file extension
    file_name += f"-{random_string(4)}{int(datetime.now().timestamp())}.xlsx";

    ymd = datetime.now().strftime("%Y/%m/%d")
    file_path = f"{file_dir}/{ymd}"

    if isinstance(sub_path, str) and not is_empty(sub_path):
        # replace multiple slashes with a single slash and clean up the path
        sub_path = sub_path.replace("//", "/").strip("/")
        file_path = f"{file_dir}/{sub_path}/{ymd}"

    os.makedirs(file_path, exist_ok=True)

    # create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = file_name.split(".")[0]
    column = []

    for key, value in column_data.items():
        width = max(len(value), 20)
        column.append({"header": value, "key": key, "width": width})

    # set the header row
    for col_idx, col in enumerate(column, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["header"])
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(
            top=Side(style="thin"), 
            left=Side(style="thin"), 
            bottom=Side(style="thin"), 
            right=Side(style="thin")
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    # add rows of data
    for row_idx, row in enumerate(row_data, start=2):
        for col_idx, col in enumerate(column, start=1):
            value = row.get(col["key"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                top=Side(style="thin"), 
                left=Side(style="thin"), 
                bottom=Side(style="thin"), 
                right=Side(style="thin")
            )
    
    # Save the workbook
    wb.save(f"{file_path}/{file_name}")

    # get file stats
    file_size = os.path.getsize(f"{file_path}/{file_name}")

    return {
        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "destination": f"{file_path}/{file_name}",
        "filename": file_name,
        "path": file_path,
        "size": file_size
    }

def read_excel(destination: str):
    result: list[dict[str, Any]] = []

    try:
        headers: list[str] = []

        if not os.path.exists(destination):
            raise ValueError("File not found")
        
        # load the workbook
        workbook = openpyxl.load_workbook(filename=destination, read_only=True)

        # process only the first worksheet
        worksheet = workbook.active

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_idx == 0:
                # process header row
                headers = [format_snakecase(cell) for cell in row if cell is not None]
            else:
                rowData: dict[str, Any] = {}

                for cell_idx, cell in enumerate(row):
                    if cell_idx < len(headers):
                        header_name = headers[cell_idx]
                        # store the cell value under the corresponding header
                        rowData[header_name] = cell

                result.append(rowData)

        return result
    except Exception:
        return result